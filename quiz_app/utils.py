import re
import json
import unicodedata
from typing import List, Dict, Any, Tuple, Optional
from openpyxl import load_workbook
from django.db import transaction
from django.conf import settings
from .models import Subject, Unit, Question
import os
import requests
from dotenv import load_dotenv

# .envファイルを読み込み
load_dotenv()


def normalize_alphanumeric(text: str) -> str:
    """
    英数字の半角・全角を正規化する
    """
    if not text:
        return text
    
    # 全角英数字を半角に変換
    normalized = unicodedata.normalize('NFKC', text)
    
    # 全角英数字のマッピング
    fullwidth_to_halfwidth = {
        '０': '0', '１': '1', '２': '2', '３': '3', '４': '4',
        '５': '5', '６': '6', '７': '7', '８': '8', '９': '9',
        'Ａ': 'A', 'Ｂ': 'B', 'Ｃ': 'C', 'Ｄ': 'D', 'Ｅ': 'E',
        'Ｆ': 'F', 'Ｇ': 'G', 'Ｈ': 'H', 'Ｉ': 'I', 'Ｊ': 'J',
        'Ｋ': 'K', 'Ｌ': 'L', 'Ｍ': 'M', 'Ｎ': 'N', 'Ｏ': 'O',
        'Ｐ': 'P', 'Ｑ': 'Q', 'Ｒ': 'R', 'Ｓ': 'S', 'Ｔ': 'T',
        'Ｕ': 'U', 'Ｖ': 'V', 'Ｗ': 'W', 'Ｘ': 'X', 'Ｙ': 'Y',
        'Ｚ': 'Z',
        'ａ': 'a', 'ｂ': 'b', 'ｃ': 'c', 'ｄ': 'd', 'ｅ': 'e',
        'ｆ': 'f', 'ｇ': 'g', 'ｈ': 'h', 'ｉ': 'i', 'ｊ': 'j',
        'ｋ': 'k', 'ｌ': 'l', 'ｍ': 'm', 'ｎ': 'n', 'ｏ': 'o',
        'ｐ': 'p', 'ｑ': 'q', 'ｒ': 'r', 'ｓ': 's', 'ｔ': 't',
        'ｕ': 'u', 'ｖ': 'v', 'ｗ': 'w', 'ｘ': 'x', 'ｙ': 'y',
        'ｚ': 'z'
    }
    
    # 全角英数字を半角に変換
    for full, half in fullwidth_to_halfwidth.items():
        normalized = normalized.replace(full, half)
    
    return normalized


def normalize_text(text: str) -> str:
    """テキストを正規化する"""
    if not text:
        return ""
    
    # 英数字の半角・全角を正規化
    text = normalize_alphanumeric(text)
    
    # 全角・半角の統一
    text = text.replace('　', ' ')  # 全角スペースを半角に
    text = text.replace('（', '(').replace('）', ')')  # 全角括弧を半角に
    text = text.replace('［', '[').replace('］', ']')  # 全角角括弧を半角に
    
    # ひらがな・カタカナの統一（カタカナに統一）
    text = text.replace('ぁ', 'ァ').replace('ぃ', 'ィ').replace('ぅ', 'ゥ').replace('ぇ', 'ェ').replace('ぉ', 'ォ')
    text = text.replace('ゃ', 'ャ').replace('ゅ', 'ュ').replace('ょ', 'ョ')
    text = text.replace('っ', 'ッ')
    text = text.replace('あ', 'ア').replace('い', 'イ').replace('う', 'ウ').replace('え', 'エ').replace('お', 'オ')
    text = text.replace('か', 'カ').replace('き', 'キ').replace('く', 'ク').replace('け', 'ケ').replace('こ', 'コ')
    text = text.replace('さ', 'サ').replace('し', 'シ').replace('す', 'ス').replace('せ', 'セ').replace('そ', 'ソ')
    text = text.replace('た', 'タ').replace('ち', 'チ').replace('つ', 'ツ').replace('て', 'テ').replace('と', 'ト')
    text = text.replace('な', 'ナ').replace('に', 'ニ').replace('ぬ', 'ヌ').replace('ね', 'ネ').replace('の', 'ノ')
    text = text.replace('は', 'ハ').replace('ひ', 'ヒ').replace('ふ', 'フ').replace('へ', 'ヘ').replace('ほ', 'ホ')
    text = text.replace('ま', 'マ').replace('み', 'ミ').replace('む', 'ム').replace('め', 'メ').replace('も', 'モ')
    text = text.replace('や', 'ヤ').replace('ゆ', 'ユ').replace('よ', 'ヨ')
    text = text.replace('ら', 'ラ').replace('り', 'リ').replace('る', 'ル').replace('れ', 'レ').replace('ろ', 'ロ')
    text = text.replace('わ', 'ワ').replace('を', 'ヲ').replace('ん', 'ン')
    
    # 前後の空白を削除
    text = text.strip()
    
    return text


def split_parts(answer: str) -> List[str]:
    """解答を「・」で分割する"""
    if not answer:
        return []
    
    parts = [part.strip() for part in answer.split('・')]
    return [part for part in parts if part]


def parse_alternatives(alternatives_text: str) -> List[str]:
    """別解テキストを解析して配列に変換"""
    if not alternatives_text:
        return []
    
    # 区切り文字で分割（/,、;など）
    alternatives = re.split(r'[/,、;；]', alternatives_text)
    return [normalize_text(alt.strip()) for alt in alternatives if alt.strip()]


def extract_unit_info(unit_text: str) -> Tuple[Optional[str], Optional[str]]:
    """単元テキストから学年とカテゴリを抽出"""
    # 例: "中1 化学" → ("中1", "化学")
    match = re.match(r'([中高][1-3])\s*(.+)', unit_text)
    if match:
        return match.group(1), match.group(2)
    return None, None


def check_answer(user_answer: str, question: Question) -> bool:
    """解答をチェックする"""
    # 正解との比較（大文字小文字、空白を無視）
    user_answer_clean = user_answer.strip().lower()
    correct_answer_clean = question.correct_answer.strip().lower()
    
    # 完全一致の場合
    if user_answer_clean == correct_answer_clean:
        return True
    
    # 別解との比較
    if question.accepted_alternatives:
        # JSONFieldの場合の対応
        if isinstance(question.accepted_alternatives, str):
            try:
                alternatives = json.loads(question.accepted_alternatives)
            except json.JSONDecodeError:
                alternatives = []
        else:
            alternatives = question.accepted_alternatives
        
        for alternative in alternatives:
            if isinstance(alternative, str):
                alternative_clean = alternative.strip().lower()
                if user_answer_clean == alternative_clean:
                    return True
    
    # 複数解答欄の場合（・で区切られている）
    if '・' in correct_answer_clean:
        correct_parts = [part.strip() for part in correct_answer_clean.split('・')]
        user_parts = [part.strip() for part in user_answer_clean.split('・')]
        
        if len(correct_parts) == len(user_parts):
            # 順序を無視して比較
            correct_parts_sorted = sorted(correct_parts)
            user_parts_sorted = sorted(user_parts)
            
            if correct_parts_sorted == user_parts_sorted:
                return True
    
    return False


def sync_alternatives_to_supabase(subject_code: str) -> Dict[str, Any]:
    """Supabaseの別解データを同期更新"""
    try:
        # 環境変数を取得
        supabase_url = os.getenv('SUPABASE_URL')
        supabase_key = os.getenv('SUPABASE_ANON_KEY')
        
        if not supabase_url or not supabase_key:
            print("❌ Supabase環境変数が設定されていません")
            return {'success': False, 'error': 'Supabase環境変数が設定されていません'}
        
        print(f"🔄 Supabase同期開始 - URL: {supabase_url}")
        
        # ヘッダーの設定
        headers = {
            'apikey': supabase_key,
            'Authorization': f'Bearer {supabase_key}',
            'Content-Type': 'application/json',
            'Prefer': 'return=representation'
        }
        
        # 対象教科の問題を取得
        subject = Subject.objects.get(code=subject_code)
        questions = Question.objects.filter(unit__subject=subject)
        
        print(f"📊 同期対象問題数: {questions.count()}件")
        
        updated_count = 0
        failed_count = 0
        errors = []
        
        for question in questions:
            try:
                # PATCHリクエストで問題を更新
                update_url = f"{supabase_url}/rest/v1/quiz_app_question?id=eq.{question.id}"
                
                # 別解データの準備
                alternatives = question.accepted_alternatives or []
                if isinstance(alternatives, str):
                    try:
                        alternatives = json.loads(alternatives)
                    except json.JSONDecodeError:
                        alternatives = []
                
                update_data = {
                    'accepted_alternatives': alternatives
                }
                
                print(f"🔄 問題ID {question.id} を更新中... 別解: {alternatives}")
                
                response = requests.patch(update_url, headers=headers, json=update_data)
                
                if response.status_code == 200:
                    updated_count += 1
                    print(f"✅ 問題ID {question.id} 更新成功")
                else:
                    failed_count += 1
                    error_msg = f"問題ID {question.id} 更新失敗 ({response.status_code}): {response.text}"
                    errors.append(error_msg)
                    print(f"❌ {error_msg}")
                    
            except Exception as e:
                failed_count += 1
                error_msg = f"問題ID {question.id} の更新エラー: {str(e)}"
                errors.append(error_msg)
                print(f"❌ {error_msg}")
        
        print(f"🎉 Supabase同期完了: {updated_count}件成功, {failed_count}件失敗")
        
        if errors:
            print("⚠️ エラー詳細:")
            for error in errors[:5]:  # 最初の5件のみ表示
                print(f"  - {error}")
        
        return {
            'success': True,
            'updated_count': updated_count,
            'failed_count': failed_count,
            'errors': errors
        }
        
    except Exception as e:
        error_msg = f"Supabase同期エラー: {str(e)}"
        print(f"❌ {error_msg}")
        return {'success': False, 'error': error_msg}


def process_xlsm_file(file_path: str, subject_code: str) -> Dict[str, Any]:
    """XLSMファイルを処理してデータを抽出"""
    try:
        workbook = load_workbook(file_path, keep_vba=True)
        worksheet = workbook.active
        
        data = []
        errors = []
        
        # ヘッダー行をスキップ（2行目から開始）
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:  # IDが空の行はスキップ
                continue
            
            try:
                # 新しい列の定義: ID, 単元, 問題, 正解, 別解, 問題タイプ, 選択肢1-6, 単位
                source_id = str(row[0]).strip()
                unit_text = str(row[1]).strip()
                question_text = str(row[2]).strip()
                correct_answer = str(row[3]).strip()
                alternatives_text = str(row[4]).strip() if row[4] else ""
                question_type = str(row[5]).strip() if len(row) > 5 and row[5] else "text"
                unit_label_text = str(row[12]).strip() if len(row) > 12 and row[12] else ""
                
                # 問題タイプの正規化
                if question_type.lower() in ['choice', '選択', '選択問題']:
                    question_type = 'choice'
                else:
                    question_type = 'text'
                
                # 選択肢の取得（G列〜L列）
                choices = []
                if question_type == 'choice':
                    # 正解を選択肢に追加
                    choices.append(correct_answer)
                    
                    # 入力された選択肢を追加
                    for i in range(6, 12):  # G列〜L列
                        if len(row) > i and row[i]:
                            choice = str(row[i]).strip()
                            if choice and choice != correct_answer:  # 重複を避ける
                                choices.append(choice)
                    
                    # 選択肢をランダムに並べ替え
                    import random
                    random.shuffle(choices)
                
                # 必須項目のチェック
                if not all([source_id, unit_text, question_text, correct_answer]):
                    errors.append(f"行{row_num}: 必須項目が不足しています")
                    continue
                
                # 単元情報の抽出
                grade, category = extract_unit_info(unit_text)
                if not grade or not category:
                    errors.append(f"行{row_num}: 単元情報の解析に失敗しました: {unit_text}")
                    continue
                
                # 別解の解析
                alternatives = parse_alternatives(alternatives_text)
                
                # 複数解答欄の判定
                parts_count = 1
                if '・' in correct_answer:
                    parts_count = len(split_parts(correct_answer))
                
                # 単位ラベルの判定
                requires_unit_label = bool(unit_label_text)
                if not unit_label_text:
                    # 単位フィールドが空の場合、問題文から自動抽出を試行
                    if any(unit in question_text.lower() for unit in ['g', 'kg', 'm', 'cm', 'l', 'ml']):
                        requires_unit_label = True
                        # 単位の抽出（簡易版）
                        unit_match = re.search(r'([0-9]+)\s*(g|kg|m|cm|l|ml)', question_text)
                        if unit_match:
                            unit_label_text = unit_match.group(2)
                
                data.append({
                    'source_id': source_id,
                    'unit_text': unit_text,
                    'grade': grade,
                    'category': category,
                    'question_text': question_text,
                    'correct_answer': correct_answer,
                    'alternatives': alternatives,
                    'question_type': question_type,
                    'choices': choices,
                    'parts_count': parts_count,
                    'requires_unit_label': requires_unit_label,
                    'unit_label_text': unit_label_text,
                })
                
            except Exception as e:
                errors.append(f"行{row_num}: 処理エラー - {str(e)}")
        
        return {
            'data': data,
            'errors': errors,
            'total_rows': len(data),
            'error_count': len(errors)
        }
        
    except Exception as e:
        return {
            'data': [],
            'errors': [f"ファイル読み込みエラー: {str(e)}"],
            'total_rows': 0,
            'error_count': 1
        }


@transaction.atomic
def save_questions_from_xlsm_data(data: List[Dict[str, Any]], subject_code: str) -> Dict[str, Any]:
    """XLSMデータから問題をデータベースに保存"""
    subject = Subject.objects.get(code=subject_code)
    saved_count = 0
    updated_count = 0
    errors = []
    
    # 対象教科の既存の別解データを完全にクリア
    print(f"🗑️ {subject.label_ja}の既存の別解データをクリア中...")
    questions_to_clear = Question.objects.filter(unit__subject=subject)
    cleared_count = 0
    for question in questions_to_clear:
        question.accepted_alternatives = []
        question.save()
        cleared_count += 1
    print(f"✅ {cleared_count}件の問題の別解データをクリアしました")
    
    for item in data:
        try:
            # 単元の取得または作成
            unit, created = Unit.objects.get_or_create(
                subject=subject,
                grade_year=item['grade'],
                category=item['category']
            )
            
            # 問題の取得または作成
            question, created = Question.objects.get_or_create(
                unit=unit,
                source_id=item['source_id'],
                defaults={
                    'question_type': item['question_type'],
                    'text': item['question_text'],
                    'correct_answer': item['correct_answer'],
                    'accepted_alternatives': item['alternatives'],
                    'choices': item['choices'],
                    'parts_count': item['parts_count'],
                    'requires_unit_label': item['requires_unit_label'],
                    'unit_label_text': item['unit_label_text'],
                }
            )
            
            if not created:
                # 既存の問題を更新（別解データは完全に上書き）
                question.question_type = item['question_type']
                question.text = item['question_text']
                question.correct_answer = item['correct_answer']
                # 別解データを完全にクリアしてから新しい別解を設定
                question.accepted_alternatives = item['alternatives'] if item['alternatives'] else []
                question.choices = item['choices']
                question.parts_count = item['parts_count']
                question.requires_unit_label = item['requires_unit_label']
                question.unit_label_text = item['unit_label_text']
                question.save()
                updated_count += 1
            else:
                saved_count += 1
                
        except Exception as e:
            errors.append(f"問題保存エラー (ID: {item['source_id']}): {str(e)}")
    
    # Supabaseとの同期
    print(f"🔄 Supabaseとの別解データ同期中...")
    sync_result = sync_alternatives_to_supabase(subject_code)
    if sync_result['success']:
        print(f"✅ Supabase同期完了: {sync_result['updated_count']}件更新, {sync_result['failed_count']}件失敗")
    else:
        print(f"⚠️ Supabase同期エラー: {sync_result['error']}")
    
    return {
        'saved_count': saved_count,
        'updated_count': updated_count,
        'errors': errors,
        'supabase_sync': sync_result
    }
